# -*_ coding:utf-8 _*-
import openpyxl

class DoExcel:
    # def __init__(self,ExcelName,SheetName):
    #              #ExcelReadName,SheetReadName,ExcelWriteName,SheetWriteName):
    #     self.ExcelName = ExcelName
    #     self.SheetName = SheetName
    # #     self.ExcelWriteName = ExcelWriteName
    # #     self.SheetWriteName = SheetWriteName
    @staticmethod
    def get_top_row(ExcelName,SheetName):
        top_row = {}
        wb = openpyxl.load_workbook(ExcelName)
        sheet = wb[SheetName]
        for i in range(1,(sheet.max_column+1)):
            top_row[sheet.cell(1,i).value] = i
        return top_row,sheet


    def read_excel(self,ExcelReadName,SheetReadName):
        """获取excel内容并返回[{},{},{}]格式"""
        excel_list = []
        top_row_read,sheet = DoExcel.get_top_row(ExcelReadName,SheetReadName)
        for i in range(2,(sheet.max_row+1)):
            each_dict = {}
            for item in top_row_read.keys():
                each_dict[item] = sheet.cell(i,top_row_read[item]).value
            excel_list.append(each_dict)
        return excel_list


    def write_excel(self,ExcelWriteName,SheetWriteName,Excel_list=[]):
        top_row_write,aa = DoExcel.get_top_row(ExcelWriteName,SheetWriteName)
        wb = openpyxl.load_workbook(ExcelWriteName)
        sheet = wb.active
        #sheet.cell(row=2, column=1, value=1)在某行某列中写入数据
        k = 2
        for item in Excel_list:
            for i in top_row_write.keys():
                sheet.cell(row=k,column=top_row_write[i],value=str(item[i]))
            k += 1
        wb.save(ExcelWriteName)
        wb.close()


if __name__ == "__main__":
    list1 = [{'TestResult': '未通过', 'res_json': {'status': 0, 'code': '20110', 'data': None, 'msg': '手机号码已被注册'}, 'real_code': '20110', 'id': 1, 'module': '注册', 'description': '正常注册'}, {'TestResult': '通过', 'res_json': {'status': 0, 'code': '20103', 'data': None, 'msg': '手机号不能为空'}, 'id': 2, 'module': '注册', 'description': '手机号为空'}, {'TestResult': '通过', 'res_json': {'status': 0, 'code': '20109', 'data': None, 'msg': '手机号码格式不正确'}, 'id': 3, 'module': '注册', 'description': '手机号位数不正确'}, {'TestResult': '通过', 'res_json': {'status': 0, 'code': '20110', 'data': None, 'msg': '手机号码已被注册'}, 'id': 4, 'module': '注册', 'description': '手机号已经被注册'}, {'TestResult': '通过', 'res_json': {'status': 0, 'code': '20103', 'data': None, 'msg': '密码不能为空'}, 'id': 5, 'module': '注册', 'description': '密码为空'}, {'TestResult': '通过', 'res_json': {'status': 0, 'code': '20108', 'data': None, 'msg': '密码长度必须为6~18'}, 'id': 6, 'module': '注册', 'description': '密码过短'}]
    DoExcel().write_excel("test_data/ExcelResult.xlsx","testCaseWrite",list1)