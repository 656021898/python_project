from openpyxl import load_workbook

# testDataXlsx = load_workbook("testData.xlsx")#创建工作表实例
# testDataSheet = testDataXlsx["Sheet1"]#获取工作表
# print(testDataSheet.cell(1,1).value)#获取某个单元格的值
# print(testDataSheet.max_column)#获取最大列数

class GetExcel:
    def __init__(self,xlsxname,sheetname):
        self.xlsxname = xlsxname
        self.sheetname = sheetname

    def get_first_line(self):
        wb = load_workbook(self.xlsxname)
        sheet = wb[self.sheetname]
        Firstline = {}
        for i in range(1,(sheet.max_column+1)):
            Firstline[sheet.cell(1,i).value] = i
        return Firstline

    def get_all_row(self):
        """根据first_line返回所有行数据，格式为[{},{}]"""
        wb = load_workbook(self.xlsxname)
        sheet = wb[self.sheetname]
        all_row = []
        for i in range(2,(sheet.max_row+1)):
            each_row = {}
            for key in self.get_first_line().keys():
                each_row[key] = sheet.cell(i,self.get_first_line()[key]).value
            all_row.append(each_row)
        return all_row


if __name__ =="__main__":
    aa = GetExcel("testData.xlsx","data").get_all_row()
    print(aa)

